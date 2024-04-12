import openpyxl

def read_data(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    row_count=sheet.max_row
    column_count=sheet.max_column
    for row in range(2,row_count+1):
        for col in range(1,column_count+1):
            print(sheet.cell(row,col).value,end=' ')
        print()

# read_data(r"D:\pytest_framework\excelfiles\loginpage.xlsx",'loginPage')
def write_data(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    sheet.cell(1,3).value='name'
    for row in range(2,5):
        for col in range(3,4):
            sheet.cell(row,col).value='Auro'
    workbook.save(filepath)
# write_data(r"D:\pytest_framework\excelfiles\loginpage.xlsx",'loginPage')


def get_data(filepath,sheetname):
    final_list=[]
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    row_count=sheet.max_row
    column_count=sheet.max_column
    for row in range(2,row_count+1):
        temp_list=[]
        for col in range(1,column_count+1):
            temp_list.append(sheet.cell(row,col).value)
        final_list.append(temp_list)
    return final_list
# get_data(r"D:\pytest_framework\excelfiles\loginpage.xlsx",'loginPage')
